import openpyxl
import os
from dotenv import load_dotenv

load_dotenv()

class ExcelUtils():
    def __init__(self, pasta:str = os.path.expanduser("~/Downloads")):
        self.pasta = pasta

    def validar_arquivo_excel(self, path):
        '''
        Função para validar se o arquivo Excel existe no caminho fornecido
        Caso não exista, levanta uma exceção.
        '''
        if not os.path.exists(path):
            raise FileNotFoundError(f"Arquivo '{path}' não existe ou não foi encontrado")

    def abrir_excel(self, path):
        '''
        Função para abrir o arquivo Excel se o caminho fornecido existir.
        Carrega a planilha ativa e retorna os dados existentes.
        '''
        try:
            self.validar_arquivo_excel(path)
            workbook = openpyxl.load_workbook(path)
            print(f"Arquivo '{path}' encontrado e carregado")
            planilha = workbook.active
            print(f"Planilha ativa: {planilha.title}")
            return planilha
        except Exception as ex:
            print(f"Erro ao abrir o arquivo Excel: {ex}")
            return None

    def dados_planilha(self, planilha):
        if planilha is None:
            print("Planilha não foi carregada corretamente.")
            return 0
        
        dados = []
        print("Dados da planilha:")
        for linha_num in range(2, planilha.max_row + 1):
            lista_dados = []
            for coluna_num in range(1, 8):
                valor_celula = planilha.cell(row=linha_num, column=coluna_num).value
                lista_dados.append(valor_celula)
            dados.append(lista_dados)
        return dados
